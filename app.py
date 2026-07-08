"""
Shopee Price Calculator — App principal (Streamlit).

Execução:
    streamlit run app.py
"""

import io
import re
import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from pricing import calcular_preco
from fees import SHOPEE_FEE_TIERS
from styles import (
    CUSTOM_CSS, TITLE_HTML, SIDEBAR_HEADER_HTML,
    section_header, badge_html,
)

# ── Configuração da página ────────────────────────────────────────────────────
st.set_page_config(
    page_title="Shopee Price Calculator",
    page_icon="🛍️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


# ── Constantes ────────────────────────────────────────────────────────────────
MARGIN_TOLERANCE = 0.02     # ±2 pp tolerância para marcar como "ok"
CURRENCY_FMT = "R$ {:,.2f}"


# ── Helpers ───────────────────────────────────────────────────────────────────
def fmt_brl(v: float) -> str:
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_pct(v: float) -> str:
    return f"{v * 100:.1f}%"


def parse_br_currency(value) -> float:
    """
    Converte texto de valor monetário para float.

    Aceita formato brasileiro (49,90 / 1.234,56), formato americano com vírgula
    de milhar (1,234.56) e strings vindas do Excel com ponto decimal (49.9),
    sem remover o ponto — que antes virava 499 incorretamente.
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if pd.isna(value):
            return 0.0
        return float(value)

    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "-"):
        return 0.0

    s = re.sub(r"[R$%\s]", "", s, flags=re.I)
    if not s:
        return 0.0

    last_comma = s.rfind(",")
    last_dot = s.rfind(".")
    has_comma = "," in s
    has_dot = "." in s

    if has_comma and has_dot:
        if last_comma > last_dot:
            # BR: vírgula é decimal; pontos são milhar (ex: 1.234,56)
            s = s.replace(".", "").replace(",", ".")
        else:
            # US: ponto decimal; vírgulas são milhar (ex: 1,234.56)
            s = s.replace(",", "")
    elif has_comma and not has_dot:
        s = s.replace(",", ".")
    elif has_dot and not has_comma:
        # Só pontos: Excel usa como decimal (49.9); BR pode usar como milhar (1.234)
        parts = s.split(".")
        if len(parts) == 2:
            frac = parts[1]
            if len(frac) <= 2:
                pass  # decimal tipo 49.90 (Excel / exportação)
            elif len(frac) == 3:
                s = parts[0] + parts[1]  # milhar BR: 1.234 → 1234
            else:
                s = "".join(parts)
        else:
            s = "".join(parts)

    try:
        return float(s)
    except ValueError:
        return 0.0


def render_cost_breakdown(row: pd.Series) -> None:
    """
    Renderiza um gráfico de pizza (donut) mostrando como o preço do produto
    se decompõe em fatias limpas:
      • Custo + Taxa Fixa Shopee  (a taxa fixa é agrupada com o custo
                                    porque é um valor fixo, não percentual)
      • Comissão Shopee %         (apenas a parte percentual)
      • Imposto + Spike Day
      • Lucro Líquido (% do preço)
      • TACOS                     (apenas se > 0)
      • Afiliado                  (apenas se > 0)
    """
    preco       = float(row["Preço Alvo (R$)"])
    custo       = float(row["Custo (R$)"])
    comissao    = float(row["Comissão Shopee (R$)"])
    taxa_fixa   = float(row["Taxa Fixa Shopee (R$)"])
    lucro       = float(row["Lucro (R$)"])
    break_even  = float(row["Break-Even (R$)"])
    margem_real = float(row["Margem Real"])
    tacos_v     = float(row.get("TACOS (R$)", 0.0) or 0.0)
    afiliado_v  = float(row.get("Afiliado (R$)", 0.0) or 0.0)

    if preco <= 0:
        return

    # Comissão Shopee % (só a parte percentual, sem a taxa fixa)
    comissao_pct = max(0.0, comissao - taxa_fixa)
    # Custo do produto + taxa fixa Shopee (agrupados na pizza)
    custo_e_taxa = custo + taxa_fixa
    # Imposto + Spike Day = o que sobra do preço após o resto.
    # Garante que os componentes somem exatamente 100%.
    imposto_spike = max(
        0.0,
        preco - comissao_pct - custo_e_taxa - lucro - tacos_v - afiliado_v,
    )

    pct_comissao = comissao_pct  / preco * 100
    pct_imposto  = imposto_spike / preco * 100
    pct_custo    = custo_e_taxa  / preco * 100
    pct_lucro    = lucro         / preco * 100
    pct_tacos    = tacos_v       / preco * 100
    pct_afiliado = afiliado_v    / preco * 100

    nome = str(row["Nome"])

    componentes = [
        ("Comissão Shopee (%)",       comissao_pct,  pct_comissao, "#E67E22"),
        ("Imposto + Spike Day",       imposto_spike, pct_imposto,  "#C0392B"),
        ("Custo + Taxa Fixa Shopee",  custo_e_taxa,  pct_custo,    "#7B2FBE"),
        ("Lucro Líquido",             lucro,         pct_lucro,    "#1DB954"),
    ]
    # Só inclui TACOS / Afiliado quando > 0 (abaixo de Lucro)
    if tacos_v > 0:
        componentes.append(
            ("TACOS (Anúncios)",      tacos_v,       pct_tacos,    "#F39C12")
        )
    if afiliado_v > 0:
        componentes.append(
            ("Comissão de Afiliado",  afiliado_v,    pct_afiliado, "#8E44AD")
        )

    # Monta as fatias da pizza usando conic-gradient (CSS puro)
    cumul = 0.0
    slices = []
    for _, _, pct, cor in componentes:
        start = cumul
        end = cumul + pct
        slices.append(f"{cor} {start:.4f}% {end:.4f}%")
        cumul = end
    conic = ", ".join(slices)

    # Legenda à direita do donut
    legenda_html = ""
    for label, valor, pct, cor in componentes:
        legenda_html += (
            f"<div style=\"display:flex; align-items:center; gap:10px; "
            f"margin:8px 0; padding:8px 12px; background:#0f0120; "
            f"border-left:3px solid {cor}; border-radius:4px;\">"
            f"<span style=\"display:inline-block; width:14px; height:14px; "
            f"background:{cor}; border-radius:3px; flex-shrink:0;\"></span>"
            f"<span style=\"flex:1; color:#e8d5ff; font-size:0.9rem;\">{label}</span>"
            f"<span style=\"color:{cor}; font-weight:700; font-size:0.92rem; "
            f"min-width:90px; text-align:right;\">{fmt_brl(valor)}</span>"
            f"<span style=\"color:#c8a8e9; font-size:0.85rem; "
            f"min-width:55px; text-align:right;\">{pct:.1f}%</span>"
            f"</div>"
        )

    html = (
        "<div style=\"background:#1f063b; border:1px solid #7B2FBE; "
        "border-radius:10px; padding:18px 20px; margin:18px 0;\">"
        "<div style=\"color:#D4AF37; font-weight:700; font-size:1rem; "
        "margin-bottom:4px;\">"
        f"🥧 Como o preço de {fmt_brl(preco)} é composto"
        "</div>"
        "<div style=\"color:#c8a8e9; font-size:0.82rem; margin-bottom:18px;\">"
        f"Primeiro produto da tabela: <b style=\"color:#e8d5ff\">{nome}</b>"
        " &nbsp;·&nbsp; Total <b style=\"color:#D4AF37\">100%</b>"
        "</div>"
        "<div style=\"display:flex; gap:32px; align-items:center; "
        "flex-wrap:wrap; justify-content:center;\">"
        f"<div style=\"width:230px; height:230px; border-radius:50%; "
        f"background:conic-gradient({conic}); position:relative; "
        "flex-shrink:0; box-shadow:0 4px 24px rgba(212, 175, 55, 0.25); "
        "border:2px solid #D4AF37;\">"
        "<div style=\"position:absolute; top:50%; left:50%; "
        "transform:translate(-50%, -50%); width:120px; height:120px; "
        "border-radius:50%; background:#1f063b; "
        "display:flex; flex-direction:column; align-items:center; "
        "justify-content:center; border:1px solid #3d1562; "
        "box-shadow:inset 0 2px 10px rgba(0, 0, 0, 0.4);\">"
        "<div style=\"font-size:0.7rem; color:#c8a8e9; "
        "text-transform:uppercase; letter-spacing:0.5px;\">Preço Alvo</div>"
        f"<div style=\"font-size:1.05rem; font-weight:700; color:#D4AF37; "
        f"margin-top:2px;\">{fmt_brl(preco)}</div>"
        "</div>"
        "</div>"
        f"<div style=\"flex:1; min-width:300px;\">{legenda_html}</div>"
        "</div>"
        # ── Caixa "Como funciona" ────────────────────────────────────────
        "<div style=\"margin-top:18px; padding:14px 16px; background:#0f0120; "
        "border-left:3px solid #1DB954; border-radius:6px; color:#c8a8e9; "
        "font-size:0.82rem; line-height:1.6;\">"
        "<div style=\"color:#1DB954; font-weight:700; font-size:0.95rem; "
        "margin-bottom:8px;\">ℹ️ Como funciona o cálculo</div>"
        # Linha 1: explicação da lógica
        "<div style=\"margin-bottom:10px;\">"
        "O cálculo é uma <b style=\"color:#D4AF37;\">regra de três</b> simples: "
        "o preço de venda é uma pizza de <b>100%</b> dividida em 4 fatias. "
        f"O <b style=\"color:#1DB954;\">Lucro Líquido</b> é "
        f"<b style=\"color:#1DB954;\">{margem_real * 100:.1f}% do preço total</b> "
        f"= <b style=\"color:#1DB954;\">{fmt_brl(lucro)}</b>, garantido por construção "
        "(é uma fatia limpa da pizza)."
        "</div>"
        # Linha 2: break-even REAL (preço onde lucro = 0)
        "<div style=\"padding:8px 10px; background:#1f063b; border-radius:4px;\">"
        "<b style=\"color:#E67E22;\">⚖️ Break-Even (preço de equilíbrio):</b> "
        f"<b style=\"color:#D4AF37;\">{fmt_brl(break_even)}</b>. "
        "Esse é o <b>preço mínimo</b> que você teria que cobrar para cobrir "
        "exatamente todas as despesas (custo do produto + comissão Shopee + "
        "imposto), <b>sem ganhar e nem perder nada</b>."
        "<div style=\"margin-top:6px; font-size:0.78rem; color:#a98ec9;\">"
        "Vender <b style=\"color:#E74C3C;\">abaixo</b> disso = prejuízo. "
        f"Vender <b style=\"color:#1DB954;\">acima</b> disso = lucro positivo. "
        f"No preço atual de <b>{fmt_brl(preco)}</b>, seu lucro líquido é "
        f"<b style=\"color:#1DB954;\">{fmt_brl(lucro)}</b> "
        f"({margem_real * 100:.1f}% do preço)."
        "</div>"
        "</div>"
        "</div>"
        "</div>"
    )
    st.markdown(html, unsafe_allow_html=True)


def parse_excel(file) -> pd.DataFrame | None:
    """
    Lê o arquivo Excel e normaliza os nomes das colunas.
    Aceita variações comuns de nomes de coluna.
    Retorna None em caso de erro.
    """
    try:
        df = pd.read_excel(file, dtype=str)
        df.columns = [str(c).strip().lower() for c in df.columns]

        col_map = {
            "id": ["id", "id_produto", "sku", "código", "codigo", "cod",
                   "mlb (item id)", "mlb(item id)", "mlb", "item id", "item_id"],
            "nome": ["nome", "nome_produto", "produto", "descrição", "descricao", "name",
                     "nome do produto"],
            "custo": ["custo", "custo_produto", "cost", "costo", "valor_custo", "custo (r$)",
                      "custo unitário", "custo unitario"],
            "peso_extra_kg": ["peso_extra_kg", "peso_extra", "peso adicional", "extra_kg"],
        }

        rename = {}
        for target, aliases in col_map.items():
            for alias in aliases:
                if alias in df.columns:
                    rename[alias] = target
                    break

        df = df.rename(columns=rename)

        if "id" not in df.columns:
            st.error("❌ Coluna de ID do produto não encontrada. "
                     "Certifique-se de ter uma coluna chamada 'ID', 'SKU' ou 'Código'.")
            return None

        if "custo" not in df.columns:
            st.error("❌ Coluna de Custo não encontrada. "
                     "Certifique-se de ter uma coluna chamada 'Custo' ou 'Cost'.")
            return None

        # Converte custo para float (BR: 49,90 / 1.234,56; Excel: 49.9)
        df["custo"] = df["custo"].map(parse_br_currency)

        if "peso_extra_kg" in df.columns:
            df["peso_extra_kg"] = df["peso_extra_kg"].map(parse_br_currency)
        else:
            df["peso_extra_kg"] = 0.0

        if "nome" not in df.columns:
            df["nome"] = df["id"].astype(str)

        df = df[df["custo"] > 0].reset_index(drop=True)
        return df

    except Exception as e:
        st.error(f"❌ Erro ao ler o arquivo: {e}")
        return None


def build_results_df(
    df_raw: pd.DataFrame,
    sobretaxa_peso: float,
    margem: float,
    desconto_promo: float,
    aliquota_imposto: float,
    tacos_pct: float = 0.0,
    afiliado_pct: float = 0.0,
    margens_override: dict[str, float] | None = None,
) -> pd.DataFrame:
    """
    Executa o motor de cálculo e retorna um DataFrame formatado.

    Se ``margens_override`` for passado (dict {id_produto: margem 0-1}),
    cada produto presente nele usa a margem específica em vez da global.
    """
    rows = []
    for p in df_raw.to_dict("records"):
        pid = str(p.get("id", ""))
        m = margem
        if margens_override and pid in margens_override:
            m = float(margens_override[pid])
        sobretaxa = sobretaxa_peso * p.get("peso_extra_kg", 0.0)
        r = calcular_preco(
            produto_id=pid,
            nome_produto=str(p.get("nome", "")),
            custo_produto=float(p.get("custo", 0.0)),
            sobretaxa_peso=sobretaxa,
            margem_desejada=m,
            desconto_promo=desconto_promo,
            aliquota_imposto=aliquota_imposto,
            tacos_pct=tacos_pct,
            afiliado_pct=afiliado_pct,
        )
        rows.append({
            "ID Produto": r.produto_id,
            "Nome": r.nome_produto,
            "Custo (R$)": r.custo,
            "Preço Alvo (R$)": r.preco_venda,
            "Break-Even (R$)": r.preco_break_even,
            "Fake Price (R$)": r.fake_price,
            "Lucro (R$)": r.lucro,
            "Comissão Shopee (R$)": r.comissao_shopee,
            "Taxa Fixa Shopee (R$)": r.taxa_fixa_shopee,
            "Imposto (R$)": r.imposto_valor,
            "TACOS (R$)": r.tacos_valor,
            "Afiliado (R$)": r.afiliado_valor,
            "Margem Real": r.margem_real,
            "Viável": r.viavel,
        })

    return pd.DataFrame(rows)


def parse_vendas_excel(file) -> pd.DataFrame | None:
    """
    Lê o relatório de Desempenho de Produtos (Parent SKU Detail) exportado
    da Shopee Seller Center — o arquivo tem várias abas (ex: "Produtos com
    Melhor Desempenho", "Impulsionar com ADS" etc.); usamos apenas a aba
    "Produtos com Melhor Desempenho", que contém a lista completa de itens
    do período. Diferente do arquivo de "Atualização em massa de produtos",
    os dados começam logo na primeira linha (sem metadados antes do cabeçalho).

    Retorna um DataFrame com: id, nome_produto, preco_efetivo — onde
    ``preco_efetivo`` é a coluna "Vendas por Pedido (Pedido Pago) (BRL)", ou
    seja, o valor médio pelo qual aquele produto foi efetivamente vendido no
    período (já líquido de qualquer promoção/cupom aplicado especificamente
    a ele — não é o preço de lista/cadastro).
    """
    try:
        try:
            df = pd.read_excel(
                file,
                sheet_name="Produtos com Melhor Desempenho",
                dtype=str,
                engine="calamine",
            )
        except ValueError:
            # Nome da aba pode variar entre exportações — usa a primeira aba.
            df = pd.read_excel(file, sheet_name=0, dtype=str, engine="calamine")

        df = df.rename(columns={
            "ID do Item":                                "id",
            "Produto":                                   "nome_produto",
            "Vendas por Pedido (Pedido Pago) (BRL)":      "preco_efetivo",
        })

        missing = [c for c in ("id", "preco_efetivo") if c not in df.columns]
        if missing:
            st.error(
                "❌ Formato inesperado do relatório de vendas. "
                "Use o relatório de Desempenho de Produtos (aba 'Produtos com "
                "Melhor Desempenho') exportado da Shopee Seller Center."
            )
            return None

        df["id"] = df["id"].astype(str).str.strip()
        df["preco_efetivo"] = df["preco_efetivo"].map(parse_br_currency)

        if "nome_produto" in df.columns:
            df["nome_produto"] = df["nome_produto"].astype(str).str.strip()
        else:
            df["nome_produto"] = df["id"]

        # Remove linhas sem ID válido ou sem venda no período
        df = df[
            df["id"].notna()
            & ~df["id"].isin(["nan", "None", ""])
            & (df["preco_efetivo"] > 0)
        ]

        # Um mesmo ID pode aparecer em mais de uma linha (ex: variação sem
        # vendas no período gera linha duplicada com valor "-" → 0.0).
        # Mantém a linha com o maior preço médio — critério conservador,
        # igual ao usado anteriormente para variações da Shopee.
        df = df.sort_values("preco_efetivo", ascending=False)
        df = df.drop_duplicates(subset="id", keep="first")

        return df[["id", "nome_produto", "preco_efetivo"]].reset_index(drop=True)

    except Exception as e:
        st.error(f"❌ Erro ao ler o relatório de vendas: {e}")
        return None


def build_diagnostico_df(
    df_raw: pd.DataFrame,
    df_vendas: pd.DataFrame,
    aliquota_imposto: float,
    afiliado_pct: float = 0.0,
    tacos_diag_pct: float = 0.0,
) -> pd.DataFrame:
    """
    Cruza a planilha de custos com o preço médio de venda efetivo (relatório
    de vendas) e calcula a margem real que cada produto está gerando hoje.

    O preço usado (``preco_efetivo``) já é o valor real de venda de cada
    produto no período (pós-promoção/cupom), então não há mais nenhum
    desconto genérico aplicado por cima dele.

    tacos_diag_pct — TACOS médio atual da conta (% sobre receita efetiva).
                      Usado sozinho no diagnóstico, independente do TACOS
                      configurado na barra lateral (não são somados).
    """
    merged = df_raw[["id", "nome", "custo"]].merge(
        df_vendas[["id", "preco_efetivo"]],
        on="id",
        how="inner",
    )

    rows = []
    for p in merged.to_dict("records"):
        custo = float(p.get("custo") or 0.0)
        P = float(p.get("preco_efetivo") or 0.0)
        if P <= 0:
            continue

        tier = SHOPEE_FEE_TIERS[0]
        for t in SHOPEE_FEE_TIERS:
            if t.price_min <= P <= t.price_max:
                tier = t
                break

        commission   = P * tier.commission_pct + tier.fixed_fee
        imposto_val  = P * aliquota_imposto
        tacos_val    = P * tacos_diag_pct
        afil_val     = P * afiliado_pct
        lucro = P - custo - commission - imposto_val - tacos_val - afil_val
        margem = lucro / P * 100 if P > 0 else 0.0

        rows.append({
            "ID Produto":                str(p.get("id", "")),
            "Nome":                      str(p.get("nome", "")),
            "Custo (R$)":                round(custo, 2),
            "Preço Médio de Venda (R$)": round(P, 2),
            "Comissão (R$)":             round(commission, 2),
            "TACOS (R$)":                round(tacos_val, 2),
            "Lucro Atual (R$)":          round(lucro, 2),
            "Margem Atual (%)":          round(margem, 2),
        })

    return pd.DataFrame(rows)


# ── Exportação Excel ──────────────────────────────────────────────────────────
def generate_excel(
    df: pd.DataFrame,
    margem: float,
    desconto_promo: float,
    aliquota: float,
    spike_day: bool = False,
    margem_individual: bool = False,
    tacos_pct: float = 0.0,
    afiliado_pct: float = 0.0,
    df_diagnostico: pd.DataFrame | None = None,
    tacos_diag_pct: float = 0.0,
) -> bytes:
    """Gera um Excel com formatação profissional e retorna bytes.

    A coluna ``Margem %`` é construída a partir de ``df["Margem Real"]`` —
    portanto reflete a margem específica de cada produto quando o usuário
    está usando o modo "Margem por linha".
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Precificação Shopee"

    # Paleta
    ROXO_ESC = "1a0533"
    ROXO_MED = "2d0a52"
    DOURADO   = "D4AF37"
    BRANCO    = "FFFFFF"

    fill_header  = PatternFill("solid", fgColor=ROXO_MED)
    fill_title   = PatternFill("solid", fgColor=ROXO_ESC)
    fill_params  = PatternFill("solid", fgColor="0f0120")
    fill_row_alt = PatternFill("solid", fgColor="1f063b")

    font_title   = Font(name="Calibri", bold=True, size=14, color=DOURADO)
    font_header  = Font(name="Calibri", bold=True, size=10, color=DOURADO)
    font_data    = Font(name="Calibri", size=10, color=BRANCO)
    font_params  = Font(name="Calibri", size=10, color="c8a8e9")

    thin_gold    = Border(
        left=Side(style="thin", color=DOURADO),
        right=Side(style="thin", color=DOURADO),
        top=Side(style="thin", color=DOURADO),
        bottom=Side(style="thin", color=DOURADO),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left", vertical="center")

    # ── Linha 1: Título ──
    ws.merge_cells("A1:J1")
    ws["A1"] = "SHOPEE PRICE CALCULATOR"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 30

    # ── Linhas 2-6: Parâmetros ──
    margem_label = (
        f"{margem*100:.1f}% (geral — ver coluna Margem %)"
        if margem_individual
        else f"{margem*100:.1f}%"
    )
    tacos_param = f"{tacos_pct * 100:.1f}%" if tacos_pct > 0 else "—"
    afiliado_param = f"{afiliado_pct * 100:.1f}%" if afiliado_pct > 0 else "—"
    params = [
        ("Margem Desejada", margem_label),
        ("Margem por linha", "Sim" if margem_individual else "Não"),
        ("Desconto Promocional (Fake Price)", f"{desconto_promo*100:.1f}%"),
        ("Imposto sobre Receita", f"{aliquota*100:.2f}%"),
        ("Spike Day", "Ativo" if spike_day else "—"),
        ("TACOS (anúncios)", tacos_param),
        ("Comissão de Afiliado", afiliado_param),
    ]
    for i, (label, value) in enumerate(params, start=2):
        ws.merge_cells(f"A{i}:E{i}")
        ws.merge_cells(f"F{i}:J{i}")
        ws[f"A{i}"] = label
        ws[f"F{i}"] = value
        ws[f"A{i}"].font = font_params
        ws[f"F{i}"].font = Font(name="Calibri", size=10, color=DOURADO, bold=True)
        ws[f"A{i}"].fill = fill_params
        ws[f"F{i}"].fill = fill_params
        ws[f"A{i}"].alignment = left_align
        ws[f"F{i}"].alignment = left_align
        ws.row_dimensions[i].height = 18

    header_row = 2 + len(params) + 1   # linha em branco antes do cabeçalho
    ws.row_dimensions[header_row - 1].height = 6   # espaço

    # ── Cabeçalhos da tabela ──
    data_cols = [
        "ID Produto", "Nome", "Custo (R$)", "Preço Alvo (R$)",
        "Fake Price (R$)", "Lucro (R$)", "Margem %",
        "Comissão Shopee (R$)", "TACOS (R$)", "Afiliado (R$)",
    ]
    for col_idx, col_name in enumerate(data_cols, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = center_align
        cell.border = thin_gold
    ws.row_dimensions[header_row].height = 22

    # ── Linhas de dados ──
    currency_num = '#,##0.00'
    pct_num      = '0.0"%"'

    first_data_row = header_row + 1
    for row_offset, (_, row) in enumerate(df.iterrows()):
        row_idx = first_data_row + row_offset
        fill_alt = fill_row_alt if row_idx % 2 == 0 else PatternFill("solid", fgColor="160430")

        margem_real_pct = float(row.get("Margem Real", 0.0)) * 100
        tacos_val = float(row.get("TACOS (R$)", 0.0) or 0.0)
        afiliado_val = float(row.get("Afiliado (R$)", 0.0) or 0.0)
        values = [
            row["ID Produto"],
            row["Nome"],
            row["Custo (R$)"],
            row["Preço Alvo (R$)"],
            row["Fake Price (R$)"],
            row["Lucro (R$)"],
            margem_real_pct,
            row["Comissão Shopee (R$)"],
            tacos_val,
            afiliado_val,
        ]
        formats = [
            None, None,
            currency_num, currency_num, currency_num,
            currency_num, pct_num, currency_num,
            currency_num, currency_num,
        ]

        for col_idx, (val, fmt) in enumerate(zip(values, formats), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill_alt
            cell.alignment = center_align if col_idx != 2 else left_align
            cell.border = Border(
                left=Side(style="thin", color="3d1562"),
                right=Side(style="thin", color="3d1562"),
                bottom=Side(style="thin", color="3d1562"),
            )
            cell.font = font_data
            if fmt:
                cell.number_format = fmt

        ws.row_dimensions[row_idx].height = 18

    # ── Largura das colunas ──
    col_widths = [14, 24, 14, 16, 16, 14, 12, 20, 14, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze panes ──
    ws.freeze_panes = f"A{first_data_row}"

    # ── Aba 2: Diagnóstico de Preços Atuais (opcional) ──────────────────────────
    if df_diagnostico is not None and len(df_diagnostico) > 0:
        ws2 = wb.create_sheet(title="Diagnóstico Atual")

        DOURADO2  = "D4AF37"
        ROXO_ESC2 = "1a0533"
        ROXO_MED2 = "2d0a52"
        BRANCO2   = "FFFFFF"
        VERMELHO2 = "E74C3C"

        fill_title2  = PatternFill("solid", fgColor=ROXO_ESC2)
        fill_header2 = PatternFill("solid", fgColor=ROXO_MED2)
        fill_params2 = PatternFill("solid", fgColor="0f0120")
        fill_normal2 = PatternFill("solid", fgColor="1f063b")
        fill_alt2    = PatternFill("solid", fgColor="160430")
        fill_neg2    = PatternFill("solid", fgColor="2e0a0a")
        fill_media2  = PatternFill("solid", fgColor="0a2010")

        font_title2  = Font(name="Calibri", bold=True, size=14, color=DOURADO2)
        font_header2 = Font(name="Calibri", bold=True, size=10, color=DOURADO2)
        font_params2 = Font(name="Calibri", size=10, color="c8a8e9")
        font_data2   = Font(name="Calibri", size=10, color=BRANCO2)
        font_neg2    = Font(name="Calibri", size=10, color=VERMELHO2, bold=True)
        font_media2  = Font(name="Calibri", bold=True, size=10, color=DOURADO2)

        thin_gold2 = Border(
            left=Side(style="thin", color=DOURADO2),
            right=Side(style="thin", color=DOURADO2),
            top=Side(style="thin", color=DOURADO2),
            bottom=Side(style="thin", color=DOURADO2),
        )
        thin_dark2 = Border(
            left=Side(style="thin", color="3d1562"),
            right=Side(style="thin", color="3d1562"),
            bottom=Side(style="thin", color="3d1562"),
        )
        ca2 = Alignment(horizontal="center", vertical="center")
        la2 = Alignment(horizontal="left", vertical="center")

        diag_cols = [
            "ID Produto", "Nome", "Custo (R$)", "Preço Médio de Venda (R$)",
            "Comissão (R$)", "TACOS (R$)",
            "Lucro Atual (R$)", "Margem Atual (%)",
        ]
        n_cols2 = len(diag_cols)
        last_col2 = get_column_letter(n_cols2)

        # ── Linha 1: Título ──
        ws2.merge_cells(f"A1:{last_col2}1")
        ws2["A1"] = "DIAGNÓSTICO DE PREÇOS ATUAIS"
        ws2["A1"].font = font_title2
        ws2["A1"].fill = fill_title2
        ws2["A1"].alignment = ca2
        ws2.row_dimensions[1].height = 30

        # ── Linhas 2-3: Parâmetros do diagnóstico ──
        diag_params = [
            ("TACOS atual da conta",  f"{tacos_diag_pct * 100:.1f}%" if tacos_diag_pct > 0 else "—"),
            ("Imposto sobre receita", f"{aliquota * 100:.2f}%"),
        ]
        half = n_cols2 // 2
        for i, (label, value) in enumerate(diag_params, start=2):
            ws2.merge_cells(f"A{i}:{get_column_letter(half)}{i}")
            ws2.merge_cells(f"{get_column_letter(half+1)}{i}:{last_col2}{i}")
            ws2[f"A{i}"] = label
            ws2[f"{get_column_letter(half+1)}{i}"] = value
            ws2[f"A{i}"].font = font_params2
            ws2[f"{get_column_letter(half+1)}{i}"].font = Font(
                name="Calibri", size=10, color=DOURADO2, bold=True
            )
            ws2[f"A{i}"].fill = fill_params2
            ws2[f"{get_column_letter(half+1)}{i}"].fill = fill_params2
            ws2[f"A{i}"].alignment = la2
            ws2[f"{get_column_letter(half+1)}{i}"].alignment = la2
            ws2.row_dimensions[i].height = 18

        # ── Linha de cabeçalho ──
        header_row = len(diag_params) + 2
        for ci, col_name in enumerate(diag_cols, start=1):
            cell = ws2.cell(row=header_row, column=ci, value=col_name)
            cell.font = font_header2
            cell.fill = fill_header2
            cell.alignment = ca2
            cell.border = thin_gold2
        ws2.row_dimensions[header_row].height = 22

        currency_num2 = '#,##0.00'
        pct_num2      = '0.00"%"'

        first_data2 = header_row + 1

        for ri, (_, row2) in enumerate(df_diagnostico.iterrows(), start=first_data2):
            margem_val = float(row2.get("Margem Atual (%)", 0.0))
            negativo = margem_val < 0
            fill_row2 = fill_neg2 if negativo else (
                fill_alt2 if ri % 2 == 0 else fill_normal2
            )

            values2 = [
                row2["ID Produto"],
                row2["Nome"],
                row2.get("Custo (R$)", 0.0),
                row2.get("Preço Médio de Venda (R$)", 0.0),
                row2.get("Comissão (R$)", 0.0),
                row2.get("TACOS (R$)", 0.0),
                row2.get("Lucro Atual (R$)", 0.0),
                margem_val,
            ]
            formats2 = [
                None, None,
                currency_num2, currency_num2,
                currency_num2, currency_num2,
                currency_num2, pct_num2,
            ]

            for ci, (val, fmt) in enumerate(zip(values2, formats2), start=1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.fill = fill_row2
                cell.alignment = ca2 if ci != 2 else la2
                cell.border = thin_dark2
                cell.font = font_neg2 if (negativo and ci == n_cols2) else font_data2
                if fmt:
                    cell.number_format = fmt
            ws2.row_dimensions[ri].height = 18

        # ── Linha de rodapé: Margem média ──
        last_data_row = first_data2 + len(df_diagnostico) - 1
        media_row = last_data_row + 1
        media_val = float(df_diagnostico["Margem Atual (%)"].mean())
        media_neg = media_val < 0
        fill_media_final = PatternFill("solid", fgColor="2e0a0a") if media_neg else fill_media2

        ws2.merge_cells(f"A{media_row}:{get_column_letter(n_cols2 - 1)}{media_row}")
        ws2[f"A{media_row}"] = "Margem Média dos Produtos"
        ws2[f"A{media_row}"].font = font_media2
        ws2[f"A{media_row}"].fill = fill_media_final
        ws2[f"A{media_row}"].alignment = la2
        ws2[f"A{media_row}"].border = thin_gold2

        media_cell = ws2.cell(row=media_row, column=n_cols2, value=media_val)
        media_cell.font = Font(
            name="Calibri", bold=True, size=11,
            color=VERMELHO2 if media_neg else "2ecc71",
        )
        media_cell.fill = fill_media_final
        media_cell.alignment = ca2
        media_cell.border = thin_gold2
        media_cell.number_format = pct_num2
        ws2.row_dimensions[media_row].height = 22

        col_widths2 = [16, 32, 13, 18, 13, 13, 15, 14]
        for i2, w2 in enumerate(col_widths2, start=1):
            ws2.column_dimensions[get_column_letter(i2)].width = w2
        ws2.freeze_panes = f"A{first_data2}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Template Excel para download ──────────────────────────────────────────────
def generate_template_excel() -> bytes:
    """Gera um Excel modelo simples para o usuário preencher e retorna os bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    ws.cell(row=1, column=1, value="MLB (Item Id)")
    ws.cell(row=1, column=2, value="Nome do Produto")
    ws.cell(row=1, column=3, value="Custo Unitário")

    example_rows = [
        ("SKU-001", "Camiseta Básica", 49.90),
        ("SKU-002", "Meia Esportiva", 18.50),
        ("SKU-003", "Tênis Casual", 110.00),
    ]
    for row_idx, (pid, nome, custo) in enumerate(example_rows, start=2):
        ws.cell(row=row_idx, column=1, value=pid)
        ws.cell(row=row_idx, column=2, value=nome)
        ws.cell(row=row_idx, column=3, value=custo)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(SIDEBAR_HEADER_HTML, unsafe_allow_html=True)

    st.markdown("**Margem de Lucro Geral**")
    margem_pct = st.slider(
        "Margem (%)", min_value=0, max_value=50, value=10, step=1,
        label_visibility="collapsed",
        help="Margem sobre o preço total de venda aplicada a todos os produtos. "
             "Ex: 10% = o lucro será 10% do preço cobrado (uma fatia limpa da pizza).",
    )

    margem_individual = st.toggle(
        "Margem específica",
        value=False,
        help="Quando ativo, a tabela passa a permitir editar a margem de cada "
             "linha individualmente. Cada linha começa com o valor da "
             "Margem de Lucro Geral acima e você ajusta apenas as que quiser. "
             "Desligar e ligar novamente restaura todas para a margem geral atual.",
    )

    st.markdown("**Desconto Fake Price**")
    promo_pct = st.slider(
        "Promo (%)", min_value=0, max_value=50, value=10, step=1,
        label_visibility="collapsed",
        help="Percentual de desconto aparente. O preço riscado será: Preço Alvo ÷ (1 - desconto%)",
    )

    st.markdown("**Imposto sobre Receita**")
    col_imposto, _ = st.columns([1, 2])
    with col_imposto:
        aliquota_input = st.text_input(
            "Imposto (%)",
            value="10,0",
            label_visibility="collapsed",
            help="Digite a alíquota de imposto sobre a receita bruta. Ex: 11,5 ou 6,73.",
        )
    aliquota_pct = max(0.0, min(50.0, parse_br_currency(aliquota_input)))
    if aliquota_input and aliquota_pct in (0.0, 50.0):
        imposto_digitado = parse_br_currency(aliquota_input)
        if imposto_digitado < 0 or imposto_digitado > 50:
            st.caption("Imposto limitado entre 0% e 50%.")

    st.markdown("**TACOS (Anúncios)**")
    tacos_pct_input = st.slider(
        "TACOS (%)", min_value=0, max_value=50, value=0, step=1,
        label_visibility="collapsed",
        help="Total Advertising Cost of Sales — percentual do preço gasto com anúncios (Shopee Ads). "
             "Deixe em 0 se não anuncia.",
    )

    st.markdown("**Comissão de Afiliado**")
    afiliado_pct_input = st.slider(
        "Afiliado (%)", min_value=0, max_value=50, value=0, step=1,
        label_visibility="collapsed",
        help="Percentual do preço pago a afiliados que indicam o produto. "
             "Deixe em 0 se não usa programa de afiliados.",
    )

    st.divider()
    st.markdown("**Taxa Spike Day**")
    spike_day = st.toggle(
        "Spike Day (+3,5%)",
        value=False,
        help="Taxa adicional cobrada pela Shopee em produtos participantes do Spike Day. Adiciona 3,5% sobre a receita bruta.",
    )
    spike_day_rate = 0.035 if spike_day else 0.0

    st.divider()
    st.markdown(
        "<div style='text-align:center; color:#7B2FBE; font-size:0.75rem;'>"
        "Taxas Shopee conforme tabela oficial<br>Versão 1.0</div>",
        unsafe_allow_html=True,
    )

# ── Parâmetros calculados ─────────────────────────────────────────────────────
margem         = margem_pct / 100
desconto_promo = promo_pct / 100
aliquota       = aliquota_pct / 100
tacos          = tacos_pct_input / 100
afiliado       = afiliado_pct_input / 100

# ── Cabeçalho principal ───────────────────────────────────────────────────────
st.markdown(TITLE_HTML, unsafe_allow_html=True)

# ── Upload ────────────────────────────────────────────────────────────────────
st.markdown(
    section_header("Importar Planilha de Produtos", "📂"),
    unsafe_allow_html=True,
)

col_upload, col_hint = st.columns([2, 3])
with col_upload:
    uploaded = st.file_uploader(
        "Selecione o arquivo Excel",
        type=["xlsx", "xls"],
        label_visibility="collapsed",
    )
    st.download_button(
        label="⬇️ Baixar Planilha Modelo",
        data=generate_template_excel(),
        file_name="modelo_shopee.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Baixe o modelo, preencha com seus produtos e faça o upload acima.",
    )
with col_hint:
    st.markdown(
        "<div style='background:#2d0a52; border:1px solid #7B2FBE; border-radius:8px;"
        " padding:12px 16px; color:#c8a8e9; font-size:0.85rem;'>"
        "<b style='color:#D4AF37'>Colunas obrigatórias:</b><br>"
        "• <code style='color:#e8d5ff'>MLB (Item Id)</code><br>"
        "• <code style='color:#e8d5ff'>Custo Unitário</code> (valor numérico em R$)<br>"
        "<b style='color:#D4AF37'>Colunas opcionais:</b><br>"
        "• <code style='color:#e8d5ff'>Nome do Produto</code>"
        "</div>",
        unsafe_allow_html=True,
    )

# ── Aviso de privacidade ──────────────────────────────────────────────────────
st.markdown(
    "<div style='background:#1f063b; border-left: 4px solid #D4AF37; "
    "border-radius:6px; padding:10px 14px; color:#c8a8e9; font-size:0.82rem; "
    "margin-top: 12px;'>"
    "🔒 <b style='color:#D4AF37'>Privacidade:</b> "
    "seus dados de custo <b>não são armazenados</b>. "
    "Eles são usados apenas para o cálculo e descartados ao atualizar a página."
    "</div>",
    unsafe_allow_html=True,
)

# ── Processamento ─────────────────────────────────────────────────────────────
if uploaded is not None:
    if "last_file" not in st.session_state or st.session_state.last_file != uploaded.name:
        st.session_state.last_file = uploaded.name
        st.session_state.df_raw = parse_excel(uploaded)
        # Ao trocar de arquivo, descarta margens individuais antigas
        st.session_state.pop("margens_override", None)

    df_raw = st.session_state.get("df_raw")

    if df_raw is not None and len(df_raw) > 0:

        # ── Margens individuais por produto ────────────────────────────────
        # Comportamento:
        #   • Toggle ligado pela primeira vez (ou após desligar): inicializa
        #     todas as linhas com o valor atual da Margem de Lucro Geral.
        #   • Toggle ligado e mantido ligado: preserva as edições do usuário;
        #     produtos novos (raro) recebem a margem geral como default.
        #   • Toggle desligado: descarta os overrides (volta a usar a margem
        #     geral para todos os produtos).
        prev_individual = st.session_state.get("margem_individual_prev", False)
        if margem_individual:
            just_turned_on = not prev_individual
            if just_turned_on or "margens_override" not in st.session_state:
                st.session_state.margens_override = {
                    str(pid): float(margem_pct)
                    for pid in df_raw["id"].astype(str)
                }
                # Bump no seed faz o data_editor remontar limpo, evitando que
                # edits antigos sejam reaplicados por cima dos novos defaults.
                st.session_state["editor_seed"] = (
                    st.session_state.get("editor_seed", 0) + 1
                )
            else:
                # Mantém edições; garante default para produtos novos
                for pid in df_raw["id"].astype(str):
                    st.session_state.margens_override.setdefault(
                        str(pid), float(margem_pct)
                    )
            margens_override = {
                pid: m / 100.0
                for pid, m in st.session_state.margens_override.items()
            }
        else:
            # Toggle desligado: descarta overrides
            st.session_state.pop("margens_override", None)
            margens_override = None
        st.session_state["margem_individual_prev"] = margem_individual

        df_results = build_results_df(
            df_raw=df_raw,
            sobretaxa_peso=0.0,
            margem=margem,
            desconto_promo=desconto_promo,
            aliquota_imposto=aliquota + spike_day_rate,
            tacos_pct=tacos,
            afiliado_pct=afiliado,
            margens_override=margens_override,
        )

        # ── Tabela de resultados ────────────────────────────────────────────
        st.markdown(
            section_header("Tabela de Preços Calculados", "📋"),
            unsafe_allow_html=True,
        )

        # Parâmetros exibidos como badges
        margem_label = (
            f"Margem {margem_pct}% (geral)" if margem_individual
            else f"Margem {margem_pct}%"
        )
        badges_html = (
            f"Parâmetros ativos: "
            + badge_html(margem_label)
            + " &nbsp; "
            + badge_html(f"Promo {promo_pct}%", "#7B2FBE")
            + " &nbsp; "
            + badge_html(f"Imposto {aliquota_pct:.1f}%", "#1a7a4a")
        )
        if spike_day:
            badges_html += " &nbsp; " + badge_html("Spike Day +3,5%", "#E67E22")
        if tacos_pct_input > 0:
            badges_html += " &nbsp; " + badge_html(f"TACOS {tacos_pct_input}%", "#C0392B")
        if afiliado_pct_input > 0:
            badges_html += " &nbsp; " + badge_html(f"Afiliado {afiliado_pct_input}%", "#8E44AD")
        if margem_individual:
            badges_html += " &nbsp; " + badge_html("Margem por linha", "#D4AF37")
        st.markdown(badges_html, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        busca_produto = st.text_input(
            "Pesquisar por ID ou Nome",
            placeholder="Digite o ID ou nome do produto...",
            help="Filtra a tabela pelo ID Produto ou pelo Nome. A exportação continua incluindo todos os produtos.",
        ).strip()
        if busca_produto:
            mask_busca = (
                df_results["ID Produto"].astype(str).str.contains(
                    busca_produto, case=False, na=False, regex=False
                )
                | df_results["Nome"].astype(str).str.contains(
                    busca_produto, case=False, na=False, regex=False
                )
            )
            df_table = df_results[mask_busca].copy()
            st.caption(
                f"{len(df_table)} de {len(df_results)} produto(s) encontrado(s)."
            )
        else:
            df_table = df_results

        # Formata o DataFrame para exibição
        cols_to_drop = [
            "Viável", "Break-Even (R$)", "Taxa Fixa Shopee (R$)",
            "Imposto (R$)", "Margem Real",
        ]
        # Só exibe TACOS / Afiliado na tabela quando o usuário tiver definido > 0
        if tacos_pct_input == 0:
            cols_to_drop.append("TACOS (R$)")
        if afiliado_pct_input == 0:
            cols_to_drop.append("Afiliado (R$)")

        display_df = df_table.drop(columns=cols_to_drop).copy()

        # ── Modo edição (margem por linha) ─────────────────────────────────
        if margem_individual:
            st.markdown(
                "<div style='color:#D4AF37; font-size:0.85rem; margin-bottom:6px;'>"
                "✏️ <b>Modo edição:</b> ajuste a coluna <b>Margem %</b> de cada "
                "linha. Os preços recalculam automaticamente."
                "</div>",
                unsafe_allow_html=True,
            )

            # Insere a coluna Margem % editável logo após o Nome
            edit_df = display_df.copy()
            margem_col = [
                float(st.session_state.margens_override[str(pid)])
                for pid in df_table["ID Produto"]
            ]
            insert_at = min(2, len(edit_df.columns))
            edit_df.insert(loc=insert_at, column="Margem %", value=margem_col)

            col_config: dict = {}
            for c in edit_df.columns:
                if c == "Margem %":
                    col_config[c] = st.column_config.NumberColumn(
                        "Margem %",
                        min_value=0,
                        max_value=50,
                        step=1,
                        format="%d%%",
                        help="Margem específica desta linha (0–50%). "
                             "Editável.",
                    )
                elif c.endswith("(R$)"):
                    col_config[c] = st.column_config.NumberColumn(
                        c,
                        format="R$ %.2f",
                        disabled=True,
                    )
                else:
                    col_config[c] = st.column_config.Column(c, disabled=True)

            # Seed na chave permite "remontar" o widget após restaurar
            # margens individuais (evita que edits antigos sejam reaplicados
            # por cima dos defaults após o reset).
            editor_seed = st.session_state.get("editor_seed", 0)
            edited = st.data_editor(
                edit_df,
                column_config=col_config,
                hide_index=True,
                use_container_width=True,
                height=420,
                key=f"editor_margens_{st.session_state.last_file}_{editor_seed}",
            )

            # Captura mudanças e dispara recálculo
            changed = False
            for pid, m in zip(df_table["ID Produto"], edited["Margem %"]):
                pid_s = str(pid)
                try:
                    new_m = float(m) if m is not None else float(margem_pct)
                except (TypeError, ValueError):
                    new_m = float(margem_pct)
                new_m = max(0.0, min(50.0, new_m))
                if (
                    abs(st.session_state.margens_override.get(pid_s, -1) - new_m)
                    > 1e-9
                ):
                    st.session_state.margens_override[pid_s] = new_m
                    changed = True

            if changed:
                st.rerun()
        else:
            # Formatação condicional por margem (cor de fundo via Styler)
            def color_rows(row):
                return ["background-color: #1f063b; color: #FFFFFF"] * len(row)

            fmt_map = {
                "Custo (R$)":          "R$ {:,.2f}",
                "Preço Alvo (R$)":     "R$ {:,.2f}",
                "Fake Price (R$)":     "R$ {:,.2f}",
                "Lucro (R$)":          "R$ {:,.2f}",
                "Comissão Shopee (R$)":"R$ {:,.2f}",
            }
            if tacos_pct_input > 0:
                fmt_map["TACOS (R$)"] = "R$ {:,.2f}"
            if afiliado_pct_input > 0:
                fmt_map["Afiliado (R$)"] = "R$ {:,.2f}"

            styled = (
                display_df.style
                .apply(color_rows, axis=1)
                .format(fmt_map)
                .set_properties(**{
                    "text-align": "center",
                    "font-size": "0.88rem",
                })
                .set_properties(subset=["Nome"], **{"text-align": "left"})
            )

            st.dataframe(styled, use_container_width=True, height=420)

        # ── Breakdown de custos do primeiro produto ────────────────────────
        if len(df_table) > 0:
            render_cost_breakdown(df_table.iloc[0])

        # ── Exportação ─────────────────────────────────────────────────────
        st.markdown(
            section_header("Exportar Resultados", "📥"),
            unsafe_allow_html=True,
        )

        excel_bytes = generate_excel(
            df=df_results,
            margem=margem,
            desconto_promo=desconto_promo,
            aliquota=aliquota,
            spike_day=spike_day,
            margem_individual=margem_individual,
            tacos_pct=tacos,
            afiliado_pct=afiliado,
        )

        col_dl, col_info = st.columns([1, 3])
        with col_dl:
            st.download_button(
                label="⬇️ Baixar Excel",
                data=excel_bytes,
                file_name="shopee_precificacao.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with col_info:
            st.markdown(
                "<div style='color:#c8a8e9; font-size:0.85rem; padding-top:8px;'>"
                "Exporta todos os produtos com os parâmetros atuais.<br>"
                "O arquivo inclui formatação profissional e destaque por status de margem."
                "</div>",
                unsafe_allow_html=True,
            )

        # ── Diagnóstico de Preços Atuais ────────────────────────────────────
        st.divider()
        st.markdown(
            section_header("Diagnóstico de Preços Atuais (Shopee)", "🔍"),
            unsafe_allow_html=True,
        )

        col_up_s, col_hint_s = st.columns([2, 3])
        with col_up_s:
            uploaded_shopee = st.file_uploader(
                "Relatório de Vendas Shopee",
                type=["xlsx", "xls"],
                key="shopee_uploader",
                label_visibility="collapsed",
            )
        with col_hint_s:
            st.markdown(
                "<div style='background:#2d0a52; border:1px solid #7B2FBE; border-radius:8px;"
                " padding:12px 16px; color:#c8a8e9; font-size:0.85rem;'>"
                "<b style='color:#D4AF37'>Como usar:</b><br>"
                "Exporte o relatório de <b>Desempenho de Produtos</b> (aba "
                "<b>Produtos com Melhor Desempenho</b>) da Shopee Seller Center "
                "e faça o upload aqui.<br>"
                "O <code style='color:#e8d5ff'>ID</code> da sua planilha de custos "
                "deve ser o <b style='color:#e8d5ff'>ID do Item</b> na Shopee.<br>"
                "A margem é calculada com base no <b style='color:#e8d5ff'>preço médio "
                "de venda real</b> de cada produto no período (coluna "
                "<i>Vendas por Pedido (Pedido Pago)</i>), já líquido de promoções e cupons."
                "</div>",
                unsafe_allow_html=True,
            )

        if uploaded_shopee is not None:
            if (
                "shopee_file" not in st.session_state
                or st.session_state.shopee_file != uploaded_shopee.name
            ):
                st.session_state.shopee_file = uploaded_shopee.name
                st.session_state.df_shopee = parse_vendas_excel(uploaded_shopee)

            df_shopee = st.session_state.get("df_shopee")

            if df_shopee is not None and len(df_shopee) > 0:

                # ── Configurações do diagnóstico ───────────────────────────
                st.markdown(
                    "<div style='background:#1a0533; border:1px solid #7B2FBE;"
                    " border-radius:8px; padding:12px 18px; margin-bottom:12px;'>"
                    "<b style='color:#D4AF37; font-size:0.9rem;'>⚙️ Parâmetros do cenário atual</b>"
                    "</div>",
                    unsafe_allow_html=True,
                )
                col_td, _ = st.columns([1, 2])
                with col_td:
                    tacos_diag_txt = st.text_input(
                        "TACOS médio atual (%)",
                        placeholder="Ex: 5",
                        key="tacos_diag_input",
                        help="% do faturamento gasto com anúncios atualmente na sua "
                             "conta Shopee (ex: 5 = 5%). Usado apenas no diagnóstico, "
                             "independente do TACOS configurado na barra lateral.",
                    )

                tacos_diag = min(max(parse_br_currency(tacos_diag_txt) / 100.0, 0.0), 1.0)

                df_diag = build_diagnostico_df(
                    df_raw=df_raw,
                    df_vendas=df_shopee,
                    aliquota_imposto=aliquota + spike_day_rate,
                    afiliado_pct=afiliado,
                    tacos_diag_pct=tacos_diag,
                )

                if len(df_diag) == 0:
                    st.warning(
                        "⚠️ Nenhum produto correspondente encontrado. "
                        "Verifique se o ID da planilha de custos coincide "
                        "com o ID do Produto na Shopee."
                    )
                else:
                    n_matched = len(df_diag)
                    n_total   = len(df_raw)
                    st.caption(
                        f"{n_matched} de {n_total} produto(s) com preço Shopee encontrado(s)."
                    )

                    # ── Filtros ──────────────────────────────────────────────
                    col_busca, col_op, col_pct = st.columns([3, 1, 1])

                    with col_busca:
                        busca_diag = st.text_input(
                            "Pesquisar por ID ou Nome",
                            placeholder="Digite o ID ou nome do produto...",
                            key="busca_diag",
                        ).strip()

                    with col_op:
                        operador = st.selectbox(
                            "Filtro de margem",
                            options=["Todos", "< menor que", "> maior que"],
                            key="op_margem_diag",
                            label_visibility="visible",
                        )

                    with col_pct:
                        filtro_pct_txt = st.text_input(
                            "Margem (%)",
                            placeholder="Ex: 10",
                            key="filtro_pct_diag",
                            disabled=(operador == "Todos"),
                            label_visibility="visible",
                        )

                    # Aplica busca por texto
                    df_diag_view = df_diag.copy()
                    if busca_diag:
                        mask_busca = (
                            df_diag_view["ID Produto"].astype(str).str.contains(
                                busca_diag, case=False, na=False, regex=False
                            )
                            | df_diag_view["Nome"].astype(str).str.contains(
                                busca_diag, case=False, na=False, regex=False
                            )
                        )
                        df_diag_view = df_diag_view[mask_busca]

                    # Aplica filtro de margem
                    filtro_margem: float | None = None
                    if operador != "Todos" and filtro_pct_txt.strip():
                        try:
                            filtro_margem = parse_br_currency(filtro_pct_txt)
                        except Exception:
                            filtro_margem = None

                    if filtro_margem is not None:
                        if operador == "< menor que":
                            df_diag_view = df_diag_view[
                                df_diag_view["Margem Atual (%)"] < filtro_margem
                            ]
                        else:
                            df_diag_view = df_diag_view[
                                df_diag_view["Margem Atual (%)"] > filtro_margem
                            ]

                    # Legenda dos filtros ativos
                    partes = []
                    if busca_diag:
                        partes.append(f'busca: "{busca_diag}"')
                    if filtro_margem is not None:
                        sinal = "<" if operador == "< menor que" else ">"
                        partes.append(f"margem {sinal} {filtro_margem:.1f}%")
                    if partes:
                        st.caption(
                            f"{len(df_diag_view)} produto(s) · filtro: {', '.join(partes)}."
                        )

                    # ── Margem média dos produtos visíveis ────────────────────
                    if len(df_diag_view) > 0:
                        media_margem = df_diag_view["Margem Atual (%)"].mean()
                        media_cor  = "#E74C3C" if media_margem < 0 else "#2ecc71"
                        media_bg   = "#2e0a0a" if media_margem < 0 else "#0a2e16"
                        media_bord = "#E74C3C" if media_margem < 0 else "#27ae60"
                        label_filtro = (
                            f" (dos {len(df_diag_view)} filtrados)"
                            if (busca_diag or filtro_margem is not None)
                            else ""
                        )
                        st.markdown(
                            f"<div style='"
                            f"display:inline-flex; align-items:center; gap:10px;"
                            f"background:{media_bg}; border:1px solid {media_bord};"
                            f"border-radius:8px; padding:10px 20px; margin-bottom:10px;"
                            f"'>"
                            f"<span style='color:#c8a8e9; font-size:0.85rem;'>"
                            f"Margem média{label_filtro}</span>"
                            f"<span style='color:{media_cor}; font-size:1.25rem;"
                            f" font-weight:700; letter-spacing:0.5px;'>"
                            f"{media_margem:.2f}%</span>"
                            f"</div>",
                            unsafe_allow_html=True,
                        )

                    def color_diag(row):
                        if row["Margem Atual (%)"] < 0:
                            return ["background-color: #2e0a0a; color: #FFFFFF"] * len(row)
                        return ["background-color: #1f063b; color: #FFFFFF"] * len(row)

                    fmt_diag = {
                        "Custo (R$)":                "R$ {:,.2f}",
                        "Preço Médio de Venda (R$)": "R$ {:,.2f}",
                        "Comissão (R$)":             "R$ {:,.2f}",
                        "TACOS (R$)":                "R$ {:,.2f}",
                        "Lucro Atual (R$)":          "R$ {:,.2f}",
                        "Margem Atual (%)":          "{:.2f}%",
                    }
                    # Oculta "TACOS" se não há tacos configurado
                    cols_drop_diag = []
                    if tacos_diag == 0.0:
                        cols_drop_diag.append("TACOS (R$)")
                    diag_display = df_diag_view.drop(
                        columns=[c for c in cols_drop_diag if c in df_diag_view.columns],
                    )

                    styled_diag = (
                        diag_display.style
                        .apply(color_diag, axis=1)
                        .format({k: v for k, v in fmt_diag.items() if k in diag_display.columns})
                        .set_properties(**{
                            "text-align": "center",
                            "font-size": "0.88rem",
                        })
                        .set_properties(subset=["Nome"], **{"text-align": "left"})
                    )
                    st.dataframe(styled_diag, use_container_width=True, height=420)

                    # Download Excel com aba de diagnóstico
                    excel_diag_bytes = generate_excel(
                        df=df_results,
                        margem=margem,
                        desconto_promo=desconto_promo,
                        aliquota=aliquota,
                        spike_day=spike_day,
                        margem_individual=margem_individual,
                        tacos_pct=tacos,
                        afiliado_pct=afiliado,
                        df_diagnostico=df_diag,
                        tacos_diag_pct=tacos_diag,
                    )
                    st.download_button(
                        label="⬇️ Baixar Excel com Diagnóstico",
                        data=excel_diag_bytes,
                        file_name="shopee_precificacao_diagnostico.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Exporta as duas abas: Precificação Shopee + Diagnóstico Atual.",
                    )

    elif df_raw is not None and len(df_raw) == 0:
        st.warning("Planilha lida, mas nenhum produto com custo > 0 foi encontrado.")

else:
    # ── Estado vazio: instruções ────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(
        """
        <div style="
            text-align: center;
            padding: 48px 24px;
            background: linear-gradient(135deg, #2d0a52 0%, #1f063b 100%);
            border: 1px dashed #7B2FBE;
            border-radius: 12px;
            color: #c8a8e9;
        ">
            <div style="font-size: 3rem; margin-bottom: 12px;">📊</div>
            <div style="font-size: 1.2rem; font-weight: 700; color: #D4AF37; margin-bottom: 8px;">
                Importe sua planilha para começar
            </div>
            <div style="font-size: 0.9rem; line-height: 1.7;">
                Faça upload de um Excel com as colunas <b>ID</b> e <b>Custo</b>.<br>
                Ajuste a margem e o desconto no painel lateral —<br>
                os preços serão calculados <b>em tempo real</b>.
            </div>
            <div style="margin-top: 20px; font-size: 0.8rem; color: #7B2FBE;">
                Taxas Shopee aplicadas automaticamente por faixa de preço
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── Tabela de referência das taxas ──────────────────────────────────────
    st.markdown(
        section_header("Tabela de Taxas Shopee (referência)", "💡"),
        unsafe_allow_html=True,
    )

    df_fees = pd.DataFrame([
        {"Faixa de Preço": "Até R$ 79,99",        "Comissão": "20% + R$ 4,00",  "Subsídio Pix": "—"},
        {"Faixa de Preço": "R$ 80 a R$ 99,99",    "Comissão": "14% + R$ 16,00", "Subsídio Pix": "5%"},
        {"Faixa de Preço": "R$ 100 a R$ 199,99",  "Comissão": "14% + R$ 20,00", "Subsídio Pix": "5%"},
        {"Faixa de Preço": "R$ 200 a R$ 499,99",  "Comissão": "14% + R$ 26,00", "Subsídio Pix": "5%"},
        {"Faixa de Preço": "Acima de R$ 500",      "Comissão": "14% + R$ 26,00", "Subsídio Pix": "8%"},
    ])
    st.dataframe(df_fees, use_container_width=True, hide_index=True)
    st.markdown(
        "<div style='font-size:0.78rem; color:#7B2FBE; margin-top:6px;'>"
        "* O Subsídio Pix é pago pela Shopee ao comprador — não afeta a receita do vendedor."
        "</div>",
        unsafe_allow_html=True,
    )
